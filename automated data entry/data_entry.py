from logging import root
from tkinter import *
from tkinter.ttk import Combobox#to import drop down menu like list
from tkinter import messagebox#to get the popups like the error or warning messages
import openpyxl,xlrd#used to read and write the excel
#xlrd to onlyread the excel files
from openpyxl import Workbook
import pathlib #When working with openpyxl to read/write Excel files, you can use pathlib to manage file paths easily.

window = Tk()
window.title("data entry")
window.geometry('700x400+300+200')
window.resizable(False,False)
window.config(bg='#326273')

file = pathlib.Path("backened_data.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1']='Full Name'
    sheet['B1']='Phone number'
    sheet['C1']='Age'
    sheet['D1']='GENDER'
    sheet['E1']='Address'
    
    file.save("backened_data.xlsx")









def Clear():
    namevalue.set('')
    contactvalue.set('')
    agevalue.set('')
    gender_combobox.set('')
    adressEntry.delete(1.0,END)

def Submit():
    name = namevalue.get()
    contact = contactvalue.get()
    age = agevalue.get()
    gender = gender_combobox.get()
    address = adressEntry.get(1.0,END)
    
    file = openpyxl.load_workbook('backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)
    file.save(r"backened_data.xlsx")
    
    messagebox.showinfo('INFO','DATA ADDED !')
    namevalue.set('')
    contactvalue.set('')
    agevalue.set('')
    gender_combobox.set('')
    adressEntry.delete(1.0,END)




#icon
# icon_img1 = PhotoImage(file="task.png")
# root.iconphoto(False,icon_img1)

#heaing
Label(window,text='plz fill out this entry form :',font='arial 13',bg="#326273",fg='#fff').place(x=20,y=20)

#label
Label(window,text='Name',font=23,bg="#326273",fg='#fff').place(x=50,y=100)
Label(window,text='contact no.',font=23,bg="#326273",fg='#fff').place(x=50,y=150)
Label(window,text='Age',font=23,bg="#326273",fg='#fff').place(x=50,y=200)
Label(window,text='Gender',font=23,bg="#326273",fg='#fff').place(x=370,y=200)
Label(window,text='Address',font=23,bg="#326273",fg='#fff').place(x=50,y=250)

#Entry
namevalue = StringVar()
contactvalue = StringVar()
agevalue = StringVar()

nameEntry = Entry(window,textvariable=namevalue,font=20,width=35,bd=2)
ContactEntry = Entry(window,textvariable=contactvalue,font=20,width=35,bd=2)
ageEntry = Entry(window,textvariable=agevalue,font=20,width=10,bd=2)


#gender
gender_combobox = Combobox(window,values=['Male','Female'],font='arial 14',state='r',width=7)
gender_combobox.place(x=450,y=200)
gender_combobox.set('Male')

#address
adressEntry = Text(window,width=50,height=4,bd=2)

adressEntry.place(x=200,y=250)
ageEntry.place(x=200,y=200)
ContactEntry.place(x=200,y=150)
nameEntry.place(x=200,y=100)

#submit
Button(window,text="Submit",bg="#326273",fg="White",width=15,height=2,activeforeground="green",command=Submit).place(x=200,y=350)
Button(window,text="Clear",bg="#326273",fg="white",width=15,height=2,activeforeground="blue",command=Clear).place(x=340,y=350)
Button(window,text="Exit",bg="#326273",fg="white",width=15,height=2,activeforeground="red",command=lambda : quit()).place(x=480,y=350)































window.mainloop()


















































